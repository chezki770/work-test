import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def load_data(input_file, sheet_name):
    """Load data from Excel."""
    return pd.read_excel(input_file, sheet_name=sheet_name)

def process_data(data):
    """Process and summarize data."""
    summary = (
        data.groupby(["תז", "שם פרטי", "שם משפחה"])["סכום"]
        .sum()
        .reset_index()
        .rename(columns={"סכום": "סה\"כ תשלום"})
    )
    return summary.sort_values(by="שם פרטי", ascending=True)

def write_output(summary, output_file):
    """Write processed data to Excel."""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="סיכום תשלומים", index=False)

def apply_styling(output_file, summary):
    """Apply conditional formatting to the Excel file."""
    workbook = load_workbook(output_file)
    worksheet = workbook["סיכום תשלומים"]
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row in range(2, len(summary) + 2):  # Skip header row
        cell_value = worksheet.cell(row=row, column=4).value  # סכום תשלום column
        if cell_value > 100:
            for col in range(1, 5):  # Apply to all columns in the row
                worksheet.cell(row=row, column=col).fill = yellow_fill
    
    workbook.save(output_file)

def main():
    input_file = "income.xlsx"
    output_file = "output.xlsx"
    sheet_name = "Sheet1"

    try:
        data = load_data(input_file, sheet_name)
        summary = process_data(data)
        write_output(summary, output_file)
        apply_styling(output_file, summary)
        print(f"The file is saved and named {output_file}")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
